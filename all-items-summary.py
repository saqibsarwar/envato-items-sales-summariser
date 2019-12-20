import sys
import csv
from openpyxl import Workbook

# input envato sales csv file
envato_sales_csv_file = sys.argv[1]

# 12.5% envato fee
envato_fee_ratio = 0.125

# python dictionary to stor themeforest item names and their summarize actual earning
envato_items = {}

# new workbook to store data from CSV file
summarized_workbook = Workbook()
sales_sheet = summarized_workbook.active

# read csv file row by row and store its data in workbook
with open( envato_sales_csv_file, 'r' ) as csv_file:
    for row in csv.reader( csv_file ):
        sales_sheet.append( row )
        # also create an item_name based index using python dictionary data structure
        item_name = row[3]
        if item_name not in envato_items:
            envato_items[item_name] = 0

# WorkSheet Cleanup
    # delete order id and order type
    sales_sheet.delete_cols(2,2)
    # remove document column
    sales_sheet.delete_cols(4,1)
    # remove other columns
    sales_sheet.delete_cols(5,12)

# fifth column of Envato Fee and sixth of Actual Earning
sales_sheet.cell( 1, 5).value = 'Envato Fee'
sales_sheet.cell( 1, 6).value = 'Actual Earning'

# starting from 2 as first row is for labels and calculate Envato Fee, Actual earning and sum that earning in end.
for row in range( 2, sales_sheet.max_row + 1 ):
    # item name from worksheet
    item_name = str( sales_sheet.cell( row, 2).value )
    # item sale price from worksheet
    price_string = sales_sheet.cell( row, 4).value
    if price_string:
        price_float = float( price_string )
        envato_fee = round( price_float * envato_fee_ratio, 2 )
        actual_earning_on_sale = price_float - envato_fee
        # assign Envato Fee to related cell in worksheet
        sales_sheet.cell( row, 5).value = envato_fee
        # assign Actual Earning to related cell in worksheet
        sales_sheet.cell( row, 6).value = actual_earning_on_sale
        # sum actual earnings in python dictionary datastructure
        if envato_items[item_name] > 0:
            envato_items[item_name] += actual_earning_on_sale
        else:
            envato_items[item_name] = actual_earning_on_sale

# create new sheet to contain summarized data of items sales.
summary_sheet = summarized_workbook.create_sheet(title='Summary')

# sort items with respec to their summarized earnings in decending order
sorted_envato_items = sorted( envato_items.items(), key = lambda kv: (kv[1], kv[0]), reverse=True )

# add summarized data to newly created sheet
for item_summary in sorted_envato_items:
    summary_sheet.append( item_summary )

# save workbook in xlsx file
summarized_workbook.save( envato_sales_csv_file.replace('.csv','-summarized.xlsx') )